import sys
import pandas as pd
from openpyxl import load_workbook

from utils.file_writer import safe_excel_write


if len(sys.argv) > 1:
    SALES_FILE_PATH = sys.argv[1]
else:
    SALES_FILE_PATH = "./input/Sales/Arjun Rao Sales JAN26.xlsx"

FINAL_OUTPUT = "./output/sales_import_ready.xlsx"

REQUIRED_COLUMNS = [
    "DATE",
    "INVOICE NO",
    "PARTICULARS",
    "GROSS VALUE"
]


def _normalize_columns(columns):
    return {
        col: " ".join(str(col).strip().upper().split())
        for col in columns
    }


def _parse_amount(value):
    if pd.isna(value):
        return 0.0

    text = str(value).replace(",", "").strip()
    if text == "":
        return 0.0

    try:
        return float(text)
    except ValueError:
        return 0.0


def _format_date(value):
    parsed = pd.to_datetime(value, errors="coerce", dayfirst=False)
    if pd.isna(parsed):
        return ""
    return parsed.strftime("%d-%m-%Y")


def _read_sales_table(sales_file_path):
    wb = load_workbook(sales_file_path, data_only=True)
    ws = wb.worksheets[0]

    if "Sales" in ws.tables:
        table = ws.tables["Sales"]
    elif ws.tables:
        first_table_name = next(iter(ws.tables))
        table = ws.tables[first_table_name]
    else:
        df = pd.read_excel(sales_file_path, sheet_name=0)
        return df

    table_data = [
        [cell.value for cell in row]
        for row in ws[table.ref]
    ]

    if not table_data:
        return pd.DataFrame()

    headers = [str(col).strip() if col is not None else "" for col in table_data[0]]
    rows = table_data[1:]
    return pd.DataFrame(rows, columns=headers)


def build_sales_vouchers(sales_file_path):
    df = _read_sales_table(sales_file_path)
    df = df.rename(columns=_normalize_columns(df.columns))

    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    records = []

    for _, row in df.iterrows():
        date = _format_date(row.get("DATE"))
        if not date:
            continue

        description = str(row.get("PARTICULARS", "")).strip()
        if description == "":
            continue

        amount = _parse_amount(row.get("GROSS VALUE"))
        if amount <= 0:
            continue

        records.append(
            {
                "Voucher_Type": "Journal",
                "Date": date,
                "Description": description,
                "Narration": str(row.get("INVOICE NO", "")).strip(),
                "Cr_Ledger": "Contract Receipts",
                "Amount": amount,
                "Cr": "CR",
                "Dr_Ledger": "S.C.Rly",
                "Dr_Amount": amount,
                "Dr": "DR"
            }
        )

    output_df = pd.DataFrame(records)
    if not output_df.empty:
        output_df.reset_index(drop=True, inplace=True)
        output_df.insert(0, "Voucher_Num", output_df.index + 1)

    return output_df


def main():
    vouchers_df = build_sales_vouchers(SALES_FILE_PATH)

    def write_vouchers():
        with pd.ExcelWriter(FINAL_OUTPUT, engine="openpyxl") as writer:
            vouchers_df.to_excel(writer, sheet_name="Journal", index=False)

    safe_excel_write(write_vouchers, FINAL_OUTPUT)

    print("\nSales import file generated successfully.")
    print(f"Rows exported: {len(vouchers_df)}")
    print(f"Output file : {FINAL_OUTPUT}")


if __name__ == "__main__":
    main()
