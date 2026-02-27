import camelot
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


STANDARD_COLUMNS = [
    "Transaction Date",
    "Value Date",
    "Description",
    "Reference Number",
    "Withdrawals",
    "Deposits",
    "Running Balance"
]


# -----------------------------------------------------
# Header Utilities
# -----------------------------------------------------

def normalize_header(text):
    return (
        str(text)
        .lower()
        .replace("\n", " ")
        .replace("/", " ")
        .strip()
    )


def map_column(col):
    col_lower = normalize_header(col)

    if "transaction" in col_lower or ("txn" in col_lower and "date" in col_lower):
        return "Transaction Date"

    if "value" in col_lower:
        return "Value Date"

    if "description" in col_lower:
        return "Description"

    if "reference" in col_lower or "cheque" in col_lower or "ref" in col_lower:
        return "Reference Number"

    if "withdraw" in col_lower or "debit" in col_lower:
        return "Withdrawals"

    if "deposit" in col_lower or "credit" in col_lower:
        return "Deposits"

    if "balance" in col_lower:
        return "Running Balance"

    return None


# -----------------------------------------------------
# Repair Merged Deposit / Balance Columns
# -----------------------------------------------------

def repair_merged_amounts(df):

    pattern = re.compile(r"^\s*([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s*$")

    for i in range(len(df)):

        deposit = str(df.loc[i, "Deposits"]).strip()
        balance = str(df.loc[i, "Running Balance"]).strip()

        # Case 1: Deposits contains both values
        match = pattern.match(deposit)
        if match and balance == "":
            df.loc[i, "Deposits"] = match.group(1)
            df.loc[i, "Running Balance"] = match.group(2)
            continue

        # Case 2: Running Balance contains both values
        match = pattern.match(balance)
        if match and deposit == "":
            df.loc[i, "Deposits"] = match.group(1)
            df.loc[i, "Running Balance"] = match.group(2)

    return df


# -----------------------------------------------------
# Merge Spillover Rows (Multiline Transactions)
# -----------------------------------------------------

def merge_spillover_rows(df):

    rows_to_drop = []

    for i in range(1, len(df)):

        txn_date = str(df.loc[i, "Transaction Date"]).strip()
        val_date = str(df.loc[i, "Value Date"]).strip()
        withdrawal = str(df.loc[i, "Withdrawals"]).strip()
        deposit = str(df.loc[i, "Deposits"]).strip()
        desc = str(df.loc[i, "Description"]).strip()

        # Spillover condition:
        if (
            desc != "" and
            txn_date == "" and
            val_date == "" and
            withdrawal == "" and
            deposit == ""
        ):
            prev_desc = str(df.loc[i - 1, "Description"]).strip()
            df.loc[i - 1, "Description"] = f"{prev_desc} {desc}"
            rows_to_drop.append(i)

    df = df.drop(index=rows_to_drop).reset_index(drop=True)

    return df


def parse_amount(value):
    if pd.isna(value):
        return 0.0

    text = str(value).replace(",", "").strip()
    if text == "" or text.lower() in {"nan", "none"}:
        return 0.0

    try:
        return float(text)
    except ValueError:
        return 0.0


def clean_text(value):
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def is_header_line(line):
    normalized = clean_text(line).upper()
    return (
        "TXN DATE" in normalized and
        "VALUE DATE" in normalized and
        "DESCRIPTION" in normalized and
        ("DEBITS" in normalized or "WITHDRAWALS" in normalized) and
        ("CREDITS" in normalized or "DEPOSITS" in normalized) and
        "BALANCE" in normalized
    )


def split_description_reference(description):
    # Best-effort reference extraction from long numeric ids in description.
    match = re.search(r"\b\d{6,}\b", description)
    reference = match.group(0) if match else ""
    return description, reference


def parse_text_transaction_line(line):
    date_regex = r"\d{2}-[A-Z]{3}-\d{4}"
    amount_regex = r"-?\d[\d,]*\.\d{2}"

    pattern = re.compile(
        rf"^({date_regex})\s+({date_regex})\s+(.*?)\s+({amount_regex})\s+({amount_regex})\s+({amount_regex})\s*$",
        re.IGNORECASE
    )

    match = pattern.match(clean_text(line))
    if not match:
        return None

    description, reference = split_description_reference(clean_text(match.group(3)))

    return {
        "Transaction Date": match.group(1).upper(),
        "Value Date": match.group(2).upper(),
        "Description": description,
        "Reference Number": reference,
        "Withdrawals": parse_amount(match.group(4)),
        "Deposits": parse_amount(match.group(5)),
        "Running Balance": parse_amount(match.group(6))
    }


def extract_text_statement(pdf_path):
    try:
        import pdfplumber
    except ImportError:
        print("pdfplumber is not installed. Text fallback unavailable.")
        return pd.DataFrame(columns=STANDARD_COLUMNS)

    transactions = []
    current_transaction = None
    date_start_pattern = re.compile(r"^\d{2}-[A-Z]{3}-\d{4}", re.IGNORECASE)

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""

            for raw_line in text.splitlines():
                line = clean_text(raw_line)

                if line == "" or is_header_line(line):
                    continue

                if date_start_pattern.match(line):
                    parsed = parse_text_transaction_line(line)
                    if not parsed:
                        continue

                    if current_transaction:
                        transactions.append(current_transaction)

                    current_transaction = parsed
                elif current_transaction:
                    # Non-date lines are treated as spillover description lines.
                    current_transaction["Description"] = clean_text(
                        f"{current_transaction['Description']} {line}"
                    )

    if current_transaction:
        transactions.append(current_transaction)

    if not transactions:
        return pd.DataFrame(columns=STANDARD_COLUMNS)

    df = pd.DataFrame(transactions)

    for column in STANDARD_COLUMNS:
        if column not in df.columns:
            df[column] = ""

    df["Description"] = df["Description"].apply(clean_text)
    df = df[STANDARD_COLUMNS]

    return df


def has_valid_transactions(df):
    required = {"Transaction Date", "Value Date", "Description", "Withdrawals", "Deposits"}
    if df.empty or not required.issubset(set(df.columns)):
        return False

    for _, row in df.iterrows():
        description = clean_text(row.get("Description", ""))
        withdrawal = parse_amount(row.get("Withdrawals"))
        deposit = parse_amount(row.get("Deposits"))

        txn_date = pd.to_datetime(row.get("Transaction Date"), errors="coerce")
        val_date = pd.to_datetime(row.get("Value Date"), errors="coerce")

        if description and (withdrawal > 0 or deposit > 0) and pd.notna(txn_date) and pd.notna(val_date):
            return True

    return False


# -----------------------------------------------------
# Main Extraction Function
# -----------------------------------------------------

def extract_bank_statement(pdf_path, output_file):

    tables = camelot.read_pdf(pdf_path, pages="all")

    print(f"Total tables detected: {tables.n}")

    dataframes = []
    accepted_tables = 0
    ignored_tables = 0

    for table in tables:

        df = table.df

        if df.empty or df.shape[0] < 2:
            ignored_tables += 1
            continue

        # First row as header
        df.columns = df.iloc[0]
        df = df[1:]

        # Dynamic column mapping
        mapped_columns = {}
        for col in df.columns:
            mapped = map_column(col)
            if mapped:
                mapped_columns[col] = mapped

        df = df.rename(columns=mapped_columns)

        # Required fields
        required = {
            "Transaction Date",
            "Value Date",
            "Description",
            "Withdrawals",
            "Deposits"
        }

        if not required.issubset(set(df.columns)):
            ignored_tables += 1
            continue

        # Add optional columns if missing
        if "Reference Number" not in df.columns:
            df["Reference Number"] = ""

        if "Running Balance" not in df.columns:
            df["Running Balance"] = ""

        # Standardize column order
        df = df[STANDARD_COLUMNS]

        accepted_tables += 1
        dataframes.append(df)

    final_df = pd.DataFrame()

    if dataframes:
        final_df = pd.concat(dataframes, ignore_index=True)

        # Clean all cell values
        for col in final_df.columns:
            final_df[col] = (
                final_df[col]
                .astype(str)
                .str.replace("\n", " ", regex=False)
                .str.replace("\r", "", regex=False)
                .str.replace(r"\s+", " ", regex=True)
                .str.strip()
            )

        # Repair merged columns
        final_df = repair_merged_amounts(final_df)

        # Merge multiline spillovers
        final_df = merge_spillover_rows(final_df)

    if not has_valid_transactions(final_df):
        print("Switching to text-based extraction (pdfplumber fallback)...")
        final_df = extract_text_statement(pdf_path)

    if not has_valid_transactions(final_df):
        raise ValueError("No valid tables found in PDF. Extraction aborted.")

    # Save Excel
    final_df.to_excel(output_file, index=False)

    # Format as Excel table
    wb = load_workbook(output_file)
    ws = wb.active

    last_row = ws.max_row
    last_col = ws.max_column
    table_range = f"A1:{ws.cell(row=last_row, column=last_col).coordinate}"

    excel_table = Table(displayName="CombinedTable", ref=table_range)

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )

    excel_table.tableStyleInfo = style
    ws.add_table(excel_table)

    wb.save(output_file)

    print(f"Tables accepted: {accepted_tables}")
    print(f"Tables ignored : {ignored_tables}")
    print("Bank statement extraction completed.")

    return final_df
